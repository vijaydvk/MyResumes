import xlrd
import xlsxwriter
import openpyxl
import string
wb = openpyxl.load_workbook('G:/vijay/Phone_order/Phone Orders 11-28-2017 Bilal file.xlsx')
ws = wb.get_sheet_by_name('Sheet1')
workbook = xlrd.open_workbook('G:/vijay/Phone_order/Phone Orders 11-28-2017 Bilal file.xlsx')
worksheet = workbook.sheet_by_name('Sheet1')
phone_code_book = xlrd.open_workbook('G:/vijay/Phone_order/Order Skus Excel.xlsx')
phone_code_sheet = phone_code_book.sheet_by_name('Skus code')
phone_code_rowcount = phone_code_sheet.nrows-1
rowcount = worksheet.nrows-1
loopcount = 1
col = 20 
ws.cell(row=loopcount, column=col, value="location code")
while loopcount <= rowcount:
	transfer_id = int(worksheet.cell(loopcount, 2).value)
	loc_code = "80193"+str(transfer_id)
	#worksheet.cell(loopcount, 1).value = transfer_id
	#worksheet.write(loopcount, 22,transfer_id)
	ws.cell(row=loopcount+1, column=col, value=loc_code)
	#print (transfer_id)
	loopcount = loopcount + 1
loopcount = 1
col = 21
while loopcount <= rowcount:
	phone_name = str(worksheet.cell(loopcount, 7).value)
	phone_name.lower()
	phone_loopcount = 1
	while phone_loopcount <= phone_code_rowcount:
		product_name = str(phone_code_sheet.cell(phone_loopcount, 0).value)
		product_name.lower()
		#print (phone_name)
		#print (product_name)
		if str(phone_name) == str(product_name):
			#print (phone_name)
			product_code = phone_code_sheet.cell(phone_loopcount, 1).value
			product_code = product_code.replace(" ","")
			#print (product_code)
			ws.cell(row=loopcount+1, column=col, value=product_code)
			#print (product_name)
			break
		phone_loopcount = phone_loopcount + 1	
	loopcount = loopcount + 1
wb.save('G:/vijay/Phone_order/Phone Orders 11-28-2017 Bilal file.xlsx')