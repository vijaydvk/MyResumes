import xlrd
import string
#import xlsxwriter
from xlrd import open_workbook
from xlwt import Workbook
from xlutils.copy import copy
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait as wait
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import NoSuchElementException
import openpyxl
def is_visible():
    try:
        wait(driver, 10).until(EC.presence_of_element_located((By.ID,'errorMessagesSub')))
        return True
    except TimeoutException:
        return False
		
def is_visible_na(val):
    try:
        driver.find_element_by_xpath('//div[contains(@id,"leftLargest")]/ul/li[2]/ul/li[3]/ul/li[4]/table/tbody/tr/td/table/tbody/tr['+str(val)+']/td[6]/span[1]')
        return True
    except:
        return False
		
def page_load_history():
    try:
        wait(driver, 1).until(EC.presence_of_element_located((By.XPATH,'//div[contains(@id,"leftLargest")]/ul/li[1]/ul/li[3]/ul/li[2]/table/tbody/tr[2]/td[1]/div/a')))
        return True
    except NoSuchElementException:
        return False

workbook = xlrd.open_workbook('G:/py/PO_OUTPUT.xls')
worksheet = workbook.sheet_by_name('Sheet1')
#-----xlutils for xls
rb = open_workbook("G:/py/PO_OUTPUT.xls")
wb = copy(rb)
s = wb.get_sheet(0)
num_rows = worksheet.nrows - 1
curr_row = 1
driver = webdriver.Chrome('C:/Users/Backoffice/Downloads/chromedriver_win32/chromedriver.exe')
driver.get("https://www.im-mobilityonline.com/bpo/secure/Login")


 # hit return after you enter search text

user_name = driver.find_element_by_id('usernameInput')
user_name.send_keys('edgar.lopez')
password = driver.find_element_by_name('password:container:input')
password.send_keys('Cricket!3')
login_button = driver.find_element_by_name(':submit')
login_button.click()
driver.get("https://www.im-mobilityonline.com/bpo/OrderHistory")
while curr_row < 10:
	print (curr_row)
	excelpo = worksheet.cell(curr_row, 0)
	excelitemcode = worksheet.cell(curr_row, 10)
	excelqty = worksheet.cell(curr_row, 7)
	potext = driver.find_element_by_name('border:containerType:widthType:border_body:purchaseOrderNumber:container:input')
	potext.send_keys(excelpo.value)
	pobutton = driver.find_element(By.XPATH, '//li[contains(@class, "fullButtonLeft")]/div/a')
	pobutton.click()
	#exception_po = wait(driver, 10).until(EC.presence_of_element_located((By.ID,'errorMessagesSub')))
	if is_visible():	
		excelpo = worksheet.cell(curr_row, 12)
		print (str(excelpo.value))
		if str(excelpo.value) == '':
			s.write(curr_row, 18, 'PO Tracking Not available')
			driver.get("https://www.im-mobilityonline.com/bpo/OrderHistory")
			curr_row += 1
		else:		
			driver.get("https://www.im-mobilityonline.com/bpo/OrderHistory")
			potext = driver.find_element_by_name('border:containerType:widthType:border_body:brightpointOrderNumber:container:input')
			potext.send_keys(excelpo.value)
			pobutton = driver.find_element(By.XPATH, '//li[contains(@class, "fullButtonLeft")]/div/a')
			pobutton.click()
			if is_visible():
				s.write(curr_row, 18, 'PO Not loading')
				driver.get("https://www.im-mobilityonline.com/bpo/OrderHistory")
				curr_row += 1
			else:
				#pohistory = driver.find_element(By.XPATH, '//div[contains(@id,"leftLargest")]/ul/li[1]/ul/li[3]/ul/li[2]/table/tbody/tr[2]/td[1]/div/a')
				if page_load_history():
					pohistory = wait(driver, 1).until(EC.presence_of_element_located((By.XPATH,'//div[contains(@id,"leftLargest")]/ul/li[1]/ul/li[3]/ul/li[2]/table/tbody/tr[2]/td[1]/div/a')))
					pohistory.click()
					#------xlutil for xls
					#s.write(curr_row, 19, pohistory.text)
					#wb.save("G:/py/PO_OUTPUT.xls")
					row_count_po_table = len(driver.find_elements_by_xpath('//div[contains(@id,"leftLargest")]/ul/li[2]/ul/li[3]/ul/li[4]/table/tbody/tr/td/table/tbody/tr'))
					#print (row_count_po_table)
					row_count_po_table = row_count_po_table - 4
					i = 2
					#print (row_count_po_table)
					while i <= row_count_po_table:
						item_code=''			
						item_code = driver.find_element_by_xpath('//div[contains(@id,"leftLargest")]/ul/li[2]/ul/li[3]/ul/li[4]/table/tbody/tr/td/table/tbody/tr['+str(i)+']/td[1]/ul/li[2]/ul[2]/li').text
						item_qty = driver.find_element_by_xpath('//div[contains(@id,"leftLargest")]/ul/li[2]/ul/li[3]/ul/li[4]/table/tbody/tr/td/table/tbody/tr['+str(i)+']/td[2]').text
						item_code_split = item_code.replace('Item Code: ', '')
						#print (item_code_split)
						#print (excelitemcode.value)
						if item_code_split == excelitemcode.value:
							print (item_qty)
							qtyval = int(excelqty.value)
							print (qtyval)			
							if int(item_qty) == qtyval:
								if is_visible_na(i):
									trackingname = driver.find_element_by_xpath('//div[contains(@id,"leftLargest")]/ul/li[2]/ul/li[3]/ul/li[4]/table/tbody/tr/td/table/tbody/tr['+str(i)+']/td[6]/span[1]').text
									trackingid = driver.find_element_by_xpath('//div[contains(@id,"leftLargest")]/ul/li[2]/ul/li[3]/ul/li[4]/table/tbody/tr/td/table/tbody/tr['+str(i)+']/td[6]/a').text
									s.write(curr_row, 13, trackingname)
									s.write(curr_row, 16, trackingid)
									wb.save("G:/py/PO_OUTPUT.xls")						
									excelponext = worksheet.cell(curr_row+1, 0)
									#print (excelpo.value)
									#print (excelponext.value)
									if excelpo.value == excelponext.value:
										excelitemcode = ''
										excelqty = ''
										excelitemcode = worksheet.cell(curr_row+1, 10)
										excelqty = worksheet.cell(curr_row+1, 7)							
										i=2
										curr_row += 1	
									else:
										driver.get("https://www.im-mobilityonline.com/bpo/OrderHistory")
										curr_row += 1	
										break
								else:
									s.write(curr_row, 13, 'NA')
									wb.save("G:/py/PO_OUTPUT.xls")
									excelponext = worksheet.cell(curr_row+1, 0)
									#print (excelpo.value)
									#print (excelponext.value)
									if excelpo.value == excelponext.value:
										excelitemcode = ''
										excelqty = ''
										excelitemcode = worksheet.cell(curr_row+1, 10)
										excelqty = worksheet.cell(curr_row+1, 7)							
										i=2
										curr_row += 1	
									else:
										driver.get("https://www.im-mobilityonline.com/bpo/OrderHistory")
										curr_row += 1	
										break
							else:
								s.write(curr_row, 13, 'QTY Not matching')				
								wb.save("G:/py/PO_OUTPUT.xls")
								driver.get("https://www.im-mobilityonline.com/bpo/OrderHistory")
								curr_row += 1
								j = i
								break
						else:
							i=i+2
					else:
						s.write(curr_row, 13, 'PO Not loading')
						driver.get("https://www.im-mobilityonline.com/bpo/OrderHistory")
						curr_row += 1
						
	else:
		#pohistory = driver.find_element(By.XPATH, '//div[contains(@id,"leftLargest")]/ul/li[1]/ul/li[3]/ul/li[2]/table/tbody/tr[2]/td[1]/div/a')
		pohistory = wait(driver, 5).until(EC.presence_of_element_located((By.XPATH,'//div[contains(@id,"leftLargest")]/ul/li[1]/ul/li[3]/ul/li[2]/table/tbody/tr[2]/td[1]/div/a')))
		pohistory.click()
		#------xlutil for xls
		#s.write(curr_row, 19, pohistory.text)
		#wb.save("G:/py/PO_OUTPUT.xls")
		row_count_po_table = len(driver.find_elements_by_xpath('//div[contains(@id,"leftLargest")]/ul/li[2]/ul/li[3]/ul/li[4]/table/tbody/tr/td/table/tbody/tr'))
		#print (row_count_po_table)
		row_count_po_table = row_count_po_table - 4
		i = 2
		#print (row_count_po_table)
		while i <= row_count_po_table:
			item_code=''			
			item_code = driver.find_element_by_xpath('//div[contains(@id,"leftLargest")]/ul/li[2]/ul/li[3]/ul/li[4]/table/tbody/tr/td/table/tbody/tr['+str(i)+']/td[1]/ul/li[2]/ul[2]/li').text
			item_qty = driver.find_element_by_xpath('//div[contains(@id,"leftLargest")]/ul/li[2]/ul/li[3]/ul/li[4]/table/tbody/tr/td/table/tbody/tr['+str(i)+']/td[2]').text
			item_code_split = item_code.replace('Item Code: ', '')
			#print (item_code_split)
			#print (excelitemcode.value)
			if item_code_split == excelitemcode.value:
				#print (item_qty)
				qtyval = int(excelqty.value)
				#print (qtyval)			
				if int(item_qty) == qtyval:
					if is_visible_na(i):
						trackingname = driver.find_element_by_xpath('//div[contains(@id,"leftLargest")]/ul/li[2]/ul/li[3]/ul/li[4]/table/tbody/tr/td/table/tbody/tr['+str(i)+']/td[6]/span[1]').text
						trackingid = driver.find_element_by_xpath('//div[contains(@id,"leftLargest")]/ul/li[2]/ul/li[3]/ul/li[4]/table/tbody/tr/td/table/tbody/tr['+str(i)+']/td[6]/a').text
						s.write(curr_row, 13, trackingname)
						s.write(curr_row, 16, trackingid)
						wb.save("G:/py/PO_OUTPUT.xls")						
						excelponext = worksheet.cell(curr_row+1, 0)
						#print (excelpo.value)
						#print (excelponext.value)
						if excelpo.value == excelponext.value:
							excelitemcode = ''
							excelqty = ''
							excelitemcode = worksheet.cell(curr_row+1, 10)
							excelqty = worksheet.cell(curr_row+1, 7)							
							i=2
							curr_row += 1	
						else:
							driver.get("https://www.im-mobilityonline.com/bpo/OrderHistory")
							curr_row += 1	
							break
					else:
						s.write(curr_row, 13, 'NA')
						wb.save("G:/py/PO_OUTPUT.xls")
						excelponext = worksheet.cell(curr_row+1, 0)
						#print (excelpo.value)
						#print (excelponext.value)
						if excelpo.value == excelponext.value:
							excelitemcode = ''
							excelqty = ''
							excelitemcode = worksheet.cell(curr_row+1, 10)
							excelqty = worksheet.cell(curr_row+1, 7)							
							i=2
							curr_row += 1	
						else:
							driver.get("https://www.im-mobilityonline.com/bpo/OrderHistory")
							curr_row += 1	
							break
				else:
					s.write(curr_row, 13, 'QTY Not matching')				
					wb.save("G:/py/PO_OUTPUT.xls")
					driver.get("https://www.im-mobilityonline.com/bpo/OrderHistory")
					curr_row += 1
					j = i
					break
			else:
				i=i+2
curr_row = 1
while curr_row < 10:
	tracking_name = worksheet.cell(curr_row, 13).value
	if "FedEx" in str(tracking_name):
		tracking_id = worksheet.cell(curr_row, 16).value
		driver.get("https://www.fedex.com/apps/fedextrack/?action=track&action=track&language=english&cntry_code=us&initial=x&tracknumbers="+str(tracking_id))
		driver.implicitly_wait(2)
		date = driver.find_element_by_xpath('//div[contains(@id,"container")]/div/div/div[2]/div/div[1]/div[2]/div[2]/div/div/div[1]/div/div[3]/div[3]/div[2]').text		
		s.write(curr_row, 13, date)
		deliver_status = driver.find_element_by_css_selector('div.statusChevron_key_status.bogus').text
		s.write(curr_row, 14, deliver_status)
		wb.save("G:/py/PO_OUTPUT.xls")		
		curr_row = curr_row + 1		
	elif "UPS" in str(tracking_name):
		curr_row = curr_row + 1
	else:
		curr_row = curr_row + 1

			