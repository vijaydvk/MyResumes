import string
import xlrd
import xlsxwriter
#import xlsxwriter
from xlrd import open_workbook
from xlwt import Workbook
from xlutils.copy import copy
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait as wait
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import StaleElementReferenceException
import openpyxl
yes = {'yes','y', 'ye', ''}

# def is_visible_captcha():
	# print ('HI')
    # try:
        # wait(driver, 40).until(EC.presence_of_element_located((By.ID,'imageText')))
        # return True
    # except TimeoutException:
        # return False
def click_to_continue():
	login_button = driver.find_element_by_id('continue')
	login_button.click()
	
def click_to_question():
	identity_button = driver.find_element(By.XPATH, '//a[contains(@id,"verify-cq-submit")]/span')
	identity_button.click()
	
def check_date(val,de):
	try:
		date = driver.find_element(By.XPATH, '/html/body/div[1]/div/div[4]/div[1]/div/table/tbody/tr['+str(val)+']/td[1]/span').text
	except (NoSuchElementException, StaleElementReferenceException):
		val = val + 1
		print (val)
		date = driver.find_element(By.XPATH, '/html/body/div[1]/div/div[4]/div[1]/div/table/tbody/tr['+str(val)+']/td[1]/span').text		
	#print (date)
	if date == de:
		return True
	else:
		return False

print ('Enter no of dates to be searched')
len_date = input()
date_input_int = 0;
date_value = [];
while date_input_int < int(len_date):
	print ("Enter date:")
	date_value.append(input())
	#date_value[date_input_int] = input()
	date_input_int = date_input_int + 1
# print ('Enter date Format like %11/07/2017%')
# date = input()
workbook = xlrd.open_workbook('G:\BOA\locationlist.xls')
worksheet = workbook.sheet_by_name('Sheet1')
rb = open_workbook("G:\BOA\locationlist.xls")
wb = copy(rb)
s = wb.get_sheet(0)
str_date = str(date_value[0])
excel_name = str.replace(str_date, '/', '-', 2)
workbook1 = xlsxwriter.Workbook('G:/py/'+str(excel_name)+'.xlsx')
worksheet1 = workbook1.add_worksheet()
sales_list = []
rowcount = worksheet.nrows-1
#print (rowcount)
driver = webdriver.Chrome('C:/Users/Backoffice/Downloads/chromedriver_win32/chromedriver.exe')
driver.get("https://www.bankofamerica.com/homepage/smallbusiness.go?page_msg=signoff&body=signoff")
user_name = driver.find_element_by_id('onlineId1')
user_name.send_keys('dan.dan')
password =  driver.find_element_by_id('passcode1')
password.send_keys('Cricket1')
login_button = driver.find_element_by_id('hp-sign-in-btn')
login_button.click()
cell_value_by_value = 0
#worksheet1.write(0, 0, 'Date')
worksheet1.write(0, 0, 'Location')
worksheet1.write(0, 1, 'Amount')
worksheet1.write(0, 2, 'Sales_Date')
newsheet = 1
print ('Too continue type y/ye/yes')
choice = input().lower()
if choice in yes:
	click_to_continue()
	while cell_value_by_value <= 10:
		url = worksheet.cell(cell_value_by_value, 3)
		loc_no = worksheet.cell(cell_value_by_value, 0)
		print (loc_no.value)
		driver.get(str(url.value))
		driver.implicitly_wait(10)
		row_count = len(driver.find_elements_by_xpath('//table[contains(@class,"transaction-records")]/tbody/tr'))
		#print (row_count)
		tr_count = 3
		while tr_count <= row_count:
			#print (tr_count)
			try:
				date = driver.find_element(By.XPATH, '/html/body/div[1]/div/div[4]/div[1]/div/table/tbody/tr['+str(tr_count)+']/td[1]/span').text
				chk_date_len = len(date_value)
				chk_date_len = chk_date_len - 1
				#print (str(date_value[chk_date_len]))
				if ( str(date) >= str(date_value[chk_date_len])):					
					date_value_int = 0						
					#print (len(date_value))				
					while date_value_int < len(date_value):
						#print (date_value_int)
						#print (date_value[date_value_int])
						if date == date_value[date_value_int]:
							print (date)
							#amount = driver.find_element(By.XPATH, '/html/body/div[1]/div/div[4]/div[1]/div/table/tbody/tr['+str(tr_count)+']/td[5]').text
							hint = driver.find_element(By.XPATH, '/html/body/div[1]/div/div[4]/div[1]/div/table/tbody/tr['+str(tr_count)+']/td[2]/div/a[1]/span[3]').text
							if "TFR TRANSFER" in hint:
								tr_count = tr_count + 1
								j = 0
							elif "MERCHANT SERVICE" in hint:
								tr_count = tr_count + 1
								j = 0
								#print (hint)
							elif "SAFE CONNECT" in hint:							
								excel_date = date.split('/')
								new_date = excel_date[1]+'-'+excel_date[0]+'-'+excel_date[2]
								#worksheet1.write(newsheet, 0, new_date)
								date_hint = hint.split("SALES DATE: ")
								new_date = date_hint[1].split(" ",1)
								#worksheet1.write(newsheet, 2, hint)
								worksheet1.write(newsheet, 0, loc_no.value)
								amount = driver.find_element(By.XPATH, '/html/body/div[1]/div/div[4]/div[1]/div/table/tbody/tr['+str(tr_count)+']/td[5]').text
								worksheet1.write(newsheet, 1, amount)
								print_date = new_date[0]
								print_date = print_date.replace('.', '')
								#print (print_date)
								comment_excel_date = print_date.split('/')
								comment_new_date = comment_excel_date[1]+'-'+comment_excel_date[0]+'-'+comment_excel_date[2]
								worksheet1.write(newsheet, 2, comment_new_date)
								if comment_new_date not in sales_list:
									sales_list.append(comment_new_date)
								#worksheet.write(newsheet, 1, cost)
								newsheet = newsheet + 1
								tr_count = tr_count + 1
							else:
								#excel_date = date.split('/')
								#new_date = excel_date[1]+'-'+excel_date[0]+'-'+excel_date[2]
								#worksheet1.write(newsheet, 0, new_date)
								worksheet1.write(newsheet, 3, hint)
								worksheet1.write(newsheet, 0, loc_no.value)
								amount = driver.find_element(By.XPATH, '/html/body/div[1]/div/div[4]/div[1]/div/table/tbody/tr['+str(tr_count)+']/td[5]').text
								worksheet1.write(newsheet, 1, amount)
								#worksheet.write(newsheet, 1, cost)
								newsheet = newsheet + 1
								tr_count = tr_count + 1
							#print (amount)
							date_value_int = date_value_int + 1
							#print (date)
						else:
							date_value_int = date_value_int + 1
				else:				
					break
			except(NoSuchElementException, StaleElementReferenceException):
				#print (tr_count)
				tr_count = tr_count + 1	
		cell_value_by_value = cell_value_by_value + 1
		#print (cell_value_by_value)
else:
   sys.stdout.write("Please respond with 'yes' or 'no'")
workbook1.close()
book = openpyxl.load_workbook('G:/py/'+str(excel_name)+'.xlsx')
sheet = book.active
rowcount = sheet.max_row
temp_row_count = rowcount + 1
sheet_no = 1
date_list_loop = 0
for i, val in enumerate(sales_list):
	book.create_sheet(val)
	new_sheet = book[str(val)]
	loop_row = 2
	data_val = 0
	data = []
	column = []
	while loop_row <= rowcount:
		op_sheet_date = sheet.cell(row=loop_row, column=3).value
		if(str(val) == str(op_sheet_date)):	
			column = []
			column.append(sheet.cell(row=loop_row, column=1).value)
			column.append(sheet.cell(row=loop_row, column=2).value)
			column.append(sheet.cell(row=loop_row, column=3).value)
			data.append(column)
		loop_row = loop_row + 1
	write_row = 0
	print (len(data))
	new_sheet.cell(row=1, column=1).value = "Location"
	new_sheet.cell(row=1, column=2).value = "Amount"
	new_sheet.cell(row=1, column=3).value = "Sales_date"
	while write_row < len(data):
		print (data[write_row][0])
		new_sheet.cell(row=write_row+2, column=1).value = data[write_row][0]
		new_sheet.cell(row=write_row+2, column=2).value = data[write_row][1]
		new_sheet.cell(row=write_row+2, column=3).value = data[write_row][2]
		write_row = write_row + 1
	print (data)
	sheet_no = sheet_no + 1
for i, val in enumerate(sales_list):
	new_sheet = book[str(val)]
	rowcount = new_sheet.max_row
	write_row_count = rowcount + 1
	write_lo_row = 2
	loc_row = 0
	while loc_row <= worksheet.nrows-1:
		loc_no = worksheet.cell(loc_row, 0).value
		while write_lo_row <= rowcount:
			new_loc_no = new_sheet.cell(row=write_lo_row, column=1).value
			if (int(loc_no) == int(new_loc_no)):
				print ('equal:'+ str(loc_no) + ' - ' + str(new_loc_no))
				present = 1
				break
			else:
				present = 0
				write_lo_row = write_lo_row + 1
		if (present == 0):
			print ('not equal:'+ str(loc_no) + ' - ' + str(new_loc_no))
			new_sheet.cell(row=write_row_count, column=1).value = loc_no
			new_sheet.cell(row=write_row_count, column=2).value = 0
			new_sheet.cell(row=write_row_count, column=3).value = str(val)
			write_row_count = write_row_count + 1
		loc_row = loc_row + 1
		write_lo_row = 2
#sheet.auto_filter.ref = 'A:C'
#sheet.auto_filter.add_filter_column(3, ['13-11-2017'] , blank=False)
#print (sheet.auto_filter.filter_column)
book.save('G:/py/'+str(excel_name)+'.xlsx')
print ("Successfully Executed")